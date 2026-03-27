"""
项目进度路由
- 支持多种项目类型的数据爬取、查询、导出
"""
import io
import time
from datetime import datetime, timezone
from typing import Optional
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query, Body, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import desc

from app.database import get_db
from app.models.progress import ProgressRecord, ProgressScrapeLog, PROJECT_TYPE_NAMES
from app.models import User, UserRole, Project, ProjectAssignment, ProjectStatus, System
from app.schemas.progress import (
    ProgressRecordResponse,
    ProgressListResponse,
    ProgressScrapeRequest,
    ProgressScrapeLogResponse,
    ProgressConfigResponse,
    ProgressConfigUpdate,
    ProgressDistributeRequest,
)
from app.services.progress_scraper import (
    ProgressScraper,
    load_config,
    save_config,
    EXCEL_COLUMNS,
    DB_FIELD_NAMES,
)
from app.services.auth import get_current_user, get_current_manager
from app.services.scrape_scheduler import scheduler

router = APIRouter(prefix="/api/progress", tags=["项目进度"])


def _validate_project_type(project_type: str) -> str:
    """校验项目类型参数"""
    if project_type not in PROJECT_TYPE_NAMES:
        raise HTTPException(
            status_code=400,
            detail=f"不支持的项目类型: {project_type}，"
                   f"可选: {', '.join(PROJECT_TYPE_NAMES.keys())}",
        )
    return project_type


@router.post("/{project_type}/scrape", response_model=ProgressScrapeLogResponse)
def scrape(
    project_type: str,
    request: Optional[ProgressScrapeRequest] = Body(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_manager),
):
    """触发数据爬取（需要经理权限）"""
    _validate_project_type(project_type)
    type_name = PROJECT_TYPE_NAMES[project_type]

    log = ProgressScrapeLog(project_type=project_type, status="running")
    db.add(log)
    db.commit()
    db.refresh(log)

    start_time = time.time()
    scraper = ProgressScraper(project_type)

    try:
        page_size = request.page_size if request else None
        records = scraper.run(page_size)

        batch_id = datetime.now().strftime("%Y%m%d_%H%M%S")

        # 增量同步：新增追加、变更更新、已删除移除
        existing = {
            r.system_id: r
            for r in db.query(ProgressRecord).filter(
                ProgressRecord.project_type == project_type
            ).all()
        }
        new_ids = set()

        for rec_data in records:
            sid = rec_data.get("system_id", "")
            new_ids.add(sid)

            if sid and sid in existing:
                # 更新已有记录
                for key, value in rec_data.items():
                    setattr(existing[sid], key, value)
                existing[sid].batch_id = batch_id
            else:
                # 新增记录
                record = ProgressRecord(
                    project_type=project_type,
                    batch_id=batch_id,
                    **rec_data,
                )
                db.add(record)

        # 删除远端已不存在的记录
        for sid, record in existing.items():
            if sid not in new_ids:
                db.delete(record)

        log.status = "success"
        log.total_records = len(records)
        log.duration_seconds = round(time.time() - start_time, 2)
        log.finished_at = datetime.now(timezone.utc)
        db.commit()
        db.refresh(log)
        return log

    except Exception as e:
        log.status = "failed"
        log.error_message = str(e)
        log.duration_seconds = round(time.time() - start_time, 2)
        log.finished_at = datetime.now(timezone.utc)
        db.commit()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"{type_name}爬取失败: {e}",
        )
    finally:
        scraper.cleanup()


@router.get("/{project_type}/records", response_model=ProgressListResponse)
def get_records(
    project_type: str,
    offset: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    search: Optional[str] = Query(None),
    batch_id: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """查询项目进度记录（分页、搜索）"""
    _validate_project_type(project_type)

    query = db.query(ProgressRecord).filter(
        ProgressRecord.project_type == project_type
    )

    # 经理只看到自己作为项目经理的记录
    if current_user.role == UserRole.manager:
        query = query.filter(
            ProgressRecord.project_manager == current_user.display_name
        )

    if batch_id:
        query = query.filter(ProgressRecord.batch_id == batch_id)

    if search:
        pattern = f"%{search}%"
        query = query.filter(
            (ProgressRecord.system_name.ilike(pattern))
            | (ProgressRecord.customer_name.ilike(pattern))
            | (ProgressRecord.project_name.ilike(pattern))
            | (ProgressRecord.project_manager.ilike(pattern))
            | (ProgressRecord.system_id.ilike(pattern))
        )

    total = query.count()
    items = query.order_by(desc(ProgressRecord.system_id)).offset(offset).limit(limit).all()

    # 批量计算分发状态：查哪些 system_id 已存在于 systems 表中
    all_system_codes = [r.system_id for r in items if r.system_id]
    distributed_codes = set()
    if all_system_codes:
        existing_systems = db.query(System.system_code).filter(
            System.system_code.in_(all_system_codes)
        ).all()
        distributed_codes = {s[0] for s in existing_systems}

    # 构造响应，手动设置 distributed 字段
    result_items = []
    for r in items:
        item = ProgressRecordResponse.model_validate(r)
        if r.system_id and r.system_id in distributed_codes:
            item.distributed = True
        result_items.append(item)

    return ProgressListResponse(total=total, items=result_items)


@router.get("/{project_type}/records/export")
def export_records(
    project_type: str,
    batch_id: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """导出项目进度数据为 Excel"""
    _validate_project_type(project_type)
    type_name = PROJECT_TYPE_NAMES[project_type]

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    query = db.query(ProgressRecord).filter(
        ProgressRecord.project_type == project_type
    )

    # 经理只导出自己作为项目经理的记录
    if current_user.role == UserRole.manager:
        query = query.filter(
            ProgressRecord.project_manager == current_user.display_name
        )

    if batch_id:
        query = query.filter(ProgressRecord.batch_id == batch_id)
    if search:
        pattern = f"%{search}%"
        query = query.filter(
            (ProgressRecord.system_name.ilike(pattern))
            | (ProgressRecord.customer_name.ilike(pattern))
            | (ProgressRecord.project_name.ilike(pattern))
            | (ProgressRecord.project_manager.ilike(pattern))
        )

    records = query.order_by(desc(ProgressRecord.id)).all()
    if not records:
        raise HTTPException(status_code=404, detail="没有可导出的数据")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = type_name

    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    data_font = Font(name="微软雅黑", size=10)
    data_align = Alignment(vertical="center", wrap_text=True)

    for col_idx, col_name in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, record in enumerate(records, 2):
        for col_idx, field in enumerate(DB_FIELD_NAMES, 1):
            value = getattr(record, field, "") or ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

    for col_idx, col_name in enumerate(EXCEL_COLUMNS, 1):
        max_len = len(col_name)
        for row_idx in range(2, min(len(records) + 2, 52)):
            val = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_len = max(max_len, len(val))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{type_name}_{timestamp}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@router.get("/{project_type}/logs", response_model=list[ProgressScrapeLogResponse])
def get_logs(
    project_type: str,
    limit: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """获取爬取日志"""
    _validate_project_type(project_type)
    logs = (
        db.query(ProgressScrapeLog)
        .filter(ProgressScrapeLog.project_type == project_type)
        .order_by(desc(ProgressScrapeLog.id))
        .limit(limit)
        .all()
    )
    return logs


@router.get("/config", response_model=ProgressConfigResponse)
def get_config(current_user=Depends(get_current_manager)):
    """获取爬虫配置（敏感字段脱敏）"""
    config = load_config()
    return ProgressConfigResponse(
        base_url=config.get("base_url", ""),
        pfx_path=config.get("pfx_path", ""),
        username=config.get("username", ""),
        has_password=bool(config.get("password")),
        has_cookie=bool(config.get("cookie")),
        page_size=config.get("page_size", 50),
    )


@router.put("/config", response_model=ProgressConfigResponse)
def update_config(
    update: ProgressConfigUpdate,
    current_user=Depends(get_current_manager),
):
    """更新爬虫配置（需要经理权限）"""
    config = load_config()
    update_data = update.model_dump(exclude_none=True)
    config.update(update_data)
    save_config(config)

    return ProgressConfigResponse(
        base_url=config.get("base_url", ""),
        pfx_path=config.get("pfx_path", ""),
        username=config.get("username", ""),
        has_password=bool(config.get("password")),
        has_cookie=bool(config.get("cookie")),
        page_size=config.get("page_size", 50),
    )


# ============ 定时爬取 ============
@router.get("/schedule/status")
def get_schedule_status(current_user=Depends(get_current_manager)):
    """获取定时爬取状态"""
    return scheduler.get_status()


@router.post("/schedule/start")
def start_schedule(
    interval_minutes: int = Query(60, ge=10, le=1440, description="间隔分钟数（10-1440）"),
    current_user=Depends(get_current_manager),
):
    """开启定时爬取"""
    scheduler.start(interval_minutes)
    return {"message": f"定时爬取已开启，间隔 {interval_minutes} 分钟", **scheduler.get_status()}


@router.post("/schedule/stop")
def stop_schedule(current_user=Depends(get_current_manager)):
    """关闭定时爬取"""
    scheduler.stop()
    return {"message": "定时爬取已关闭", **scheduler.get_status()}



def _parse_project_code(raw_code: str) -> str:
    """项目编号：去掉系统编号末尾的子系统序号（如 QZXGC-202602007-03 → QZXGC-202602007）"""
    parts = raw_code.rsplit("-", 1)
    if len(parts) == 2 and parts[1].isdigit() and len(parts[1]) <= 2:
        return parts[0]
    return raw_code


@router.post("/records/{record_id}/distribute")
def distribute_record(
    record_id: int,
    request: ProgressDistributeRequest,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_manager),
):
    """从爬取记录快速创建项目并分发给员工（需要经理权限）
    - 项目编号不存在时：创建新项目 + 系统 + 分配员工
    - 项目编号已存在时：追加系统到已有项目 + 补充分配未分配的员工
    """
    record = db.query(ProgressRecord).filter(ProgressRecord.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")

    # 校验员工ID
    employees = db.query(User).filter(
        User.id.in_(request.assignee_ids),
        User.role == UserRole.employee,
    ).all()
    if len(employees) != len(request.assignee_ids):
        raise HTTPException(status_code=400, detail="部分员工ID无效或不是员工角色")

    # 业务类别映射
    type_name = PROJECT_TYPE_NAMES.get(record.project_type, "等保测评")

    raw_code = record.system_id or f"AUTO_{record.id}"
    project_code = _parse_project_code(raw_code)

    # 检查项目是否已存在
    existing = db.query(Project).filter(
        Project.project_code == project_code
    ).first()

    if existing:
        # 项目已存在 → 检查该系统是否已添加
        if record.system_id:
            dup_system = db.query(System).filter(
                System.project_id == existing.id,
                System.system_code == record.system_id,
            ).first()
            if dup_system:
                raise HTTPException(
                    status_code=400,
                    detail=f"系统 {record.system_id} 已分发到项目 {project_code}",
                )

        # 追加系统到已有项目
        if record.system_name:
            system = System(
                project_id=existing.id,
                system_code=record.system_id,
                system_name=record.system_name,
                system_level=record.system_level or "第二级",
            )
            db.add(system)

        # 补充分配未分配的员工
        new_count = 0
        for emp in employees:
            already = db.query(ProjectAssignment).filter(
                ProjectAssignment.project_id == existing.id,
                ProjectAssignment.assignee_id == emp.id,
            ).first()
            if not already:
                db.add(ProjectAssignment(
                    project_id=existing.id,
                    assignee_id=emp.id,
                ))
                new_count += 1

        db.commit()
        return {
            "message": f"已将系统追加到项目 {project_code}" + (f"，新增分配 {new_count} 名员工" if new_count else ""),
            "project_id": existing.id,
            "project_code": project_code,
        }

    # 项目不存在 → 新建项目
    project = Project(
        project_code=project_code,
        project_name=record.project_name or record.system_name or "未命名项目",
        client_name=record.customer_name,
        business_category=type_name,
        project_location=record.project_location,
        contract_status=record.contract_status or "未签订",
        filing_status=record.register_status or "未备案",
        business_manager_name=record.sale_contact,
        implementation_manager_id=employees[0].id,
        creator_id=current_user.id,
        status=ProjectStatus.assigned,
        remark=request.remark,
        contact_name=record.contact_name,
        contact_phone=record.contact_phone,
    )
    db.add(project)
    db.flush()

    # 创建系统信息
    if record.system_name:
        system = System(
            project_id=project.id,
            system_code=record.system_id,
            system_name=record.system_name,
            system_level=record.system_level or "第二级",
        )
        db.add(system)

    # 分配给选中的员工
    for emp in employees:
        assignment = ProjectAssignment(
            project_id=project.id,
            assignee_id=emp.id,
        )
        db.add(assignment)

    db.commit()

    return {
        "message": f"已创建项目并分发给 {len(employees)} 名员工",
        "project_id": project.id,
        "project_code": project.project_code,
    }
